"""
Export Manager - Excel Export with Comprehensive Formatting

Purpose: Export selected Q&A pairs to professional Excel format
Compliance: Business deliverable for $8/video revenue model
Architecture: Multi-sheet Excel with formatting, metrics, and insights

Excel Structure:
- Sheet 1: Selected Questions (Top 4 Q&A pairs)
- Sheet 2: All Candidates (All 30 generated questions)
- Sheet 3: Metrics & Insights (Performance summary)
- Sheet 4: Patterns (Learned patterns)
"""

import logging
from typing import List, Dict, Optional, Any
from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Export format options"""
    EXCEL = "excel"          # .xlsx format
    CSV = "csv"              # .csv format
    JSON = "json"            # .json format


@dataclass
class ExportConfig:
    """Configuration for export"""
    
    # Format options
    default_format: ExportFormat = ExportFormat.EXCEL
    
    # Excel styling
    use_colors: bool = True
    use_bold_headers: bool = True
    auto_adjust_columns: bool = True
    freeze_panes: bool = True
    
    # Content options
    include_evidence_details: bool = True
    include_timestamps: bool = True
    include_gemini_answers: bool = True
    include_metrics_sheet: bool = True
    include_patterns_sheet: bool = True
    
    # File naming
    include_video_id: bool = True
    include_timestamp: bool = True
    
    def get_filename(
        self,
        video_id: Optional[str] = None,
        base_name: str = "adversarial_qa"
    ) -> str:
        """Generate filename based on config"""
        parts = [base_name]
        
        if self.include_video_id and video_id:
            parts.append(video_id)
        
        if self.include_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            parts.append(timestamp)
        
        filename = "_".join(parts)
        
        if self.default_format == ExportFormat.EXCEL:
            return f"{filename}.xlsx"
        elif self.default_format == ExportFormat.CSV:
            return f"{filename}.csv"
        elif self.default_format == ExportFormat.JSON:
            return f"{filename}.json"
        
        return filename


class ExcelExporter:
    """
    Excel exporter with professional formatting
    
    Requires: openpyxl library
    """
    
    def __init__(self, config: Optional[ExportConfig] = None):
        """
        Initialize Excel exporter
        
        Args:
            config: Export configuration
        """
        self.config = config or ExportConfig()
        
        # Try to import openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            self.openpyxl = openpyxl
            self.Font = Font
            self.PatternFill = PatternFill
            self.Alignment = Alignment
            self.Border = Border
            self.Side = Side
            self.get_column_letter = get_column_letter
            
            self.available = True
        except ImportError:
            logger.warning("openpyxl not available - Excel export will be limited")
            self.available = False
    
    def export(
        self,
        selected_questions: List[Dict[str, Any]],
        all_questions: Optional[List[Dict[str, Any]]] = None,
        metrics: Optional[Dict[str, Any]] = None,
        patterns: Optional[Dict[str, Any]] = None,
        video_id: Optional[str] = None,
        output_path: Optional[Path] = None
    ) -> Path:
        """
        Export to Excel with multiple sheets
        
        Args:
            selected_questions: Top 4 selected questions
            all_questions: All generated questions (optional)
            metrics: Performance metrics (optional)
            patterns: Learned patterns (optional)
            video_id: Video identifier
            output_path: Output file path
            
        Returns:
            Path to exported file
        """
        if not self.available:
            raise ImportError("openpyxl is required for Excel export. "
                            "Install with: pip install openpyxl")
        
        logger.info(f"Exporting {len(selected_questions)} selected questions to Excel")
        
        # Create workbook
        wb = self.openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Sheet 1: Selected Questions (Top 4)
        self._create_selected_sheet(wb, selected_questions, video_id)
        
        # Sheet 2: All Candidates (if provided)
        if all_questions:
            self._create_all_candidates_sheet(wb, all_questions)
        
        # Sheet 3: Metrics & Insights (if provided)
        if self.config.include_metrics_sheet and metrics:
            self._create_metrics_sheet(wb, metrics)
        
        # Sheet 4: Patterns (if provided)
        if self.config.include_patterns_sheet and patterns:
            self._create_patterns_sheet(wb, patterns)
        
        # Determine output path
        if not output_path:
            filename = self.config.get_filename(video_id)
            output_path = Path(filename)
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Excel exported to {output_path}")
        
        return output_path
    
    def _create_selected_sheet(
        self,
        wb,
        selected_questions: List[Dict[str, Any]],
        video_id: Optional[str]
    ) -> None:
        """Create sheet with selected top 4 questions"""
        
        ws = wb.create_sheet("Selected Questions", 0)
        
        # Title row
        ws['A1'] = f"Top 4 Adversarial Questions - Video {video_id or 'Unknown'}"
        ws['A1'].font = self.Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = [
            'Rank', 'Question Type', 'Question', 'Golden Answer',
            'Gemini Answer', 'Hallucination', 'Difficulty', 'Reason'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.Font(bold=True, color='FFFFFF')
            cell.fill = self.PatternFill(start_color='4472C4', 
                                         end_color='4472C4', 
                                         fill_type='solid')
            cell.alignment = self.Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for idx, q in enumerate(selected_questions, start=1):
            row = idx + 3
            
            ws.cell(row=row, column=1, value=q.get('selection_rank', idx))
            ws.cell(row=row, column=2, value=q.get('question_type', ''))
            ws.cell(row=row, column=3, value=q.get('question', ''))
            ws.cell(row=row, column=4, value=q.get('answer', ''))
            ws.cell(row=row, column=5, value=q.get('gemini_answer', ''))
            ws.cell(row=row, column=6, value=q.get('hallucination_type', 'none'))
            ws.cell(row=row, column=7, value=q.get('difficulty_score', 0))
            ws.cell(row=row, column=8, value=q.get('selection_reason', ''))
            
            # Wrap text for readability
            for col in [3, 4, 5, 8]:
                ws.cell(row=row, column=col).alignment = self.Alignment(
                    wrap_text=True, vertical='top'
                )
            
            # Color code hallucination severity
            hall_cell = ws.cell(row=row, column=6)
            hall_type = q.get('hallucination_type', 'none').lower()
            
            if hall_type == 'critical':
                hall_cell.fill = self.PatternFill(start_color='FF0000',
                                                   end_color='FF0000',
                                                   fill_type='solid')
                hall_cell.font = self.Font(color='FFFFFF', bold=True)
            elif hall_type == 'major':
                hall_cell.fill = self.PatternFill(start_color='FFC000',
                                                   end_color='FFC000',
                                                   fill_type='solid')
            elif hall_type == 'minor':
                hall_cell.fill = self.PatternFill(start_color='FFFF00',
                                                   end_color='FFFF00',
                                                   fill_type='solid')
        
        # Auto-adjust column widths
        if self.config.auto_adjust_columns:
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 50
        
        # Freeze panes
        if self.config.freeze_panes:
            ws.freeze_panes = 'A4'
    
    def _create_all_candidates_sheet(
        self,
        wb,
        all_questions: List[Dict[str, Any]]
    ) -> None:
        """Create sheet with all candidate questions"""
        
        ws = wb.create_sheet("All Candidates")
        
        # Title
        ws['A1'] = "All Generated Questions"
        ws['A1'].font = self.Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ['ID', 'Type', 'Question', 'Answer', 'Tier', 'Validation Status']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.Font(bold=True, color='FFFFFF')
            cell.fill = self.PatternFill(start_color='70AD47',
                                         end_color='70AD47',
                                         fill_type='solid')
            cell.alignment = self.Alignment(horizontal='center')
        
        # Data rows
        for idx, q in enumerate(all_questions, start=1):
            row = idx + 3
            
            ws.cell(row=row, column=1, value=q.get('question_id', ''))
            ws.cell(row=row, column=2, value=q.get('question_type', ''))
            ws.cell(row=row, column=3, value=q.get('question', ''))
            ws.cell(row=row, column=4, value=q.get('answer', ''))
            ws.cell(row=row, column=5, value=q.get('tier', ''))
            ws.cell(row=row, column=6, value=q.get('validation_status', 'unknown'))
            
            # Wrap text
            for col in [3, 4]:
                ws.cell(row=row, column=col).alignment = self.Alignment(
                    wrap_text=True, vertical='top'
                )
        
        # Auto-adjust column widths
        if self.config.auto_adjust_columns:
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 20
        
        # Freeze panes
        if self.config.freeze_panes:
            ws.freeze_panes = 'A4'
    
    def _create_metrics_sheet(
        self,
        wb,
        metrics: Dict[str, Any]
    ) -> None:
        """Create sheet with performance metrics"""
        
        ws = wb.create_sheet("Metrics & Insights")
        
        # Title
        ws['A1'] = "Performance Metrics & Insights"
        ws['A1'].font = self.Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        row = 3
        
        # Overall metrics
        ws.cell(row=row, column=1, value="Overall Metrics").font = self.Font(bold=True)
        row += 1
        
        overall_metrics = [
            ('Total Questions Generated', metrics.get('total_questions', 0)),
            ('Validation Pass Rate', f"{metrics.get('validation_pass_rate', 0):.1%}"),
            ('Gemini Fail Rate', f"{metrics.get('gemini_fail_rate', 0):.1%}"),
            ('Hallucination Rate', f"{metrics.get('hallucination_rate', 0):.2%}"),
            ('Questions Selected', metrics.get('selected_count', 4)),
        ]
        
        for metric_name, metric_value in overall_metrics:
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=2, value=metric_value)
            row += 1
        
        row += 1
        
        # Diversity metrics
        diversity = metrics.get('diversity_metrics', {})
        if diversity:
            ws.cell(row=row, column=1, value="Diversity Metrics").font = self.Font(bold=True)
            row += 1
            
            diversity_metrics = [
                ('Diversity Score', f"{diversity.get('diversity_score', 0):.3f}"),
                ('Type Coverage', f"{diversity.get('coverage_ratio', 0):.1%}"),
                ('Unique Types', diversity.get('type_distribution', {}).get('unique_types', 0)),
                ('Redundancy Count', diversity.get('redundancy_count', 0)),
            ]
            
            for metric_name, metric_value in diversity_metrics:
                ws.cell(row=row, column=1, value=metric_name)
                ws.cell(row=row, column=2, value=metric_value)
                row += 1
        
        row += 1
        
        # Target achievement
        ws.cell(row=row, column=1, value="Target Achievement").font = self.Font(bold=True)
        row += 1
        
        targets_met = metrics.get('targets_met', {})
        target_rows = [
            ('Validation Target (90%)', '✓' if targets_met.get('validation', False) else '✗'),
            ('Gemini Failures Target (30%)', '✓' if targets_met.get('gemini_failures', False) else '✗'),
            ('Hallucination Target (<0.1%)', '✓' if targets_met.get('hallucinations', False) else '✗'),
        ]
        
        for target_name, achieved in target_rows:
            ws.cell(row=row, column=1, value=target_name)
            cell = ws.cell(row=row, column=2, value=achieved)
            
            if achieved == '✓':
                cell.font = self.Font(color='008000', bold=True)
            else:
                cell.font = self.Font(color='FF0000', bold=True)
            
            row += 1
        
        # Auto-adjust columns
        if self.config.auto_adjust_columns:
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 20
    
    def _create_patterns_sheet(
        self,
        wb,
        patterns: Dict[str, Any]
    ) -> None:
        """Create sheet with learned patterns"""
        
        ws = wb.create_sheet("Learned Patterns")
        
        # Title
        ws['A1'] = "Learned Patterns & Recommendations"
        ws['A1'].font = self.Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Recommendations
        recommendations = patterns.get('recommendations', [])
        if recommendations:
            ws.cell(row=row, column=1, value="Recommendations").font = self.Font(bold=True)
            row += 1
            
            for rec in recommendations:
                ws.cell(row=row, column=1, value=f"• {rec}")
                ws.cell(row=row, column=1).alignment = self.Alignment(wrap_text=True)
                row += 1
            
            row += 1
        
        # Best question types
        best_types = patterns.get('best_question_types', [])
        if best_types:
            ws.cell(row=row, column=1, value="Top Performing Question Types").font = self.Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value="Type")
            ws.cell(row=row, column=2, value="Success Rate")
            
            for cell in [ws.cell(row=row, column=1), ws.cell(row=row, column=2)]:
                cell.font = self.Font(bold=True)
                cell.fill = self.PatternFill(start_color='D9E1F2',
                                             end_color='D9E1F2',
                                             fill_type='solid')
            row += 1
            
            for type_info in best_types:
                ws.cell(row=row, column=1, value=type_info['type'])
                ws.cell(row=row, column=2, value=f"{type_info['success_rate']:.1%}")
                row += 1
            
            row += 1
        
        # Detected patterns
        pattern_list = patterns.get('patterns', [])
        if pattern_list:
            ws.cell(row=row, column=1, value="Detected Patterns").font = self.Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value="Pattern")
            ws.cell(row=row, column=2, value="Success Rate")
            ws.cell(row=row, column=3, value="Confidence")
            ws.cell(row=row, column=4, value="Occurrences")
            
            for col in [1, 2, 3, 4]:
                cell = ws.cell(row=row, column=col)
                cell.font = self.Font(bold=True)
                cell.fill = self.PatternFill(start_color='D9E1F2',
                                             end_color='D9E1F2',
                                             fill_type='solid')
            row += 1
            
            for pattern in pattern_list[:10]:  # Top 10 patterns
                ws.cell(row=row, column=1, value=pattern.get('description', ''))
                ws.cell(row=row, column=2, value=f"{pattern.get('success_rate', 0):.1%}")
                ws.cell(row=row, column=3, value=f"{pattern.get('confidence', 0):.1%}")
                ws.cell(row=row, column=4, value=pattern.get('occurrences', 0))
                
                ws.cell(row=row, column=1).alignment = self.Alignment(wrap_text=True)
                row += 1
        
        # Auto-adjust columns
        if self.config.auto_adjust_columns:
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15


class ExportManager:
    """
    Manages export of Q&A data to various formats
    
    Primary format: Excel with multiple sheets
    Alternative formats: CSV, JSON
    """
    
    def __init__(self, config: Optional[ExportConfig] = None):
        """
        Initialize export manager
        
        Args:
            config: Export configuration
        """
        self.config = config or ExportConfig()
        self.excel_exporter = ExcelExporter(config=self.config)
        
        logger.info("ExportManager initialized")
        logger.info(f"Default format: {self.config.default_format.value}")
    
    def export_to_excel(
        self,
        video_id: str,
        selected_questions: List[Dict[str, Any]],
        all_questions: Optional[List[Dict[str, Any]]] = None,
        feedback_result: Optional[Dict[str, Any]] = None,
        learning_insights: Optional[Dict[str, Any]] = None,
        output_path: Optional[Path] = None
    ) -> Path:
        """
        Export to Excel format (primary deliverable)
        
        Args:
            video_id: Video identifier
            selected_questions: Top 4 selected questions
            all_questions: All generated questions
            feedback_result: Feedback metrics
            learning_insights: Learned patterns and insights
            output_path: Output file path
            
        Returns:
            Path to exported Excel file
        """
        logger.info(f"Exporting results for video {video_id} to Excel")
        
        return self.excel_exporter.export(
            selected_questions=selected_questions,
            all_questions=all_questions,
            metrics=feedback_result,
            patterns=learning_insights,
            video_id=video_id,
            output_path=output_path
        )
    
    def export_to_json(
        self,
        data: Dict[str, Any],
        output_path: Path
    ) -> Path:
        """
        Export to JSON format
        
        Args:
            data: Data to export
            output_path: Output file path
            
        Returns:
            Path to exported JSON file
        """
        import json
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
        
        logger.info(f"Exported to JSON: {output_path}")
        
        return output_path
    
    def export_to_csv(
        self,
        selected_questions: List[Dict[str, Any]],
        output_path: Path
    ) -> Path:
        """
        Export to CSV format (simple)
        
        Args:
            selected_questions: Questions to export
            output_path: Output file path
            
        Returns:
            Path to exported CSV file
        """
        import csv
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Define CSV columns
        columns = [
            'rank', 'question_type', 'question', 'answer',
            'gemini_answer', 'hallucination_type', 'difficulty_score',
            'selection_reason'
        ]
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            
            for q in selected_questions:
                writer.writerow({
                    'rank': q.get('selection_rank', 0),
                    'question_type': q.get('question_type', ''),
                    'question': q.get('question', ''),
                    'answer': q.get('answer', ''),
                    'gemini_answer': q.get('gemini_answer', ''),
                    'hallucination_type': q.get('hallucination_type', ''),
                    'difficulty_score': q.get('difficulty_score', 0),
                    'selection_reason': q.get('selection_reason', '')
                })
        
        logger.info(f"Exported to CSV: {output_path}")
        
        return output_path
